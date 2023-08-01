import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from scipy.signal import find_peaks
import mplcursors
from matplotlib.widgets import Cursor
from matplotlib.lines import Line2D
import openpyxl
from smoothing_functions import *

# importar el nombre del archivo modificado
with open('names_data.py') as f:
    code = compile(f.read(), 'names_data.py', 'exec')
    exec(code)

# importar la data a un dataframe
# name_data = 'C140C_new.txt'
data = pd.read_table('01_Data_pre_procesada/' + name_data, sep = '\t', names = ['step', 'date', 'ch0', 'ch1', 'ch2', 'ch3', 'ch4', 'cabezal', 'nucleo', 'esf', 'def'], skiprows = 3, header = None, encoding = 'latin')

# convertir a listas las fuerzas, los desplazamientos y pasos
P = np.array(data['ch0'])
D = np.array(data['nucleo'])

# índice donde se encuentra la máxima carga
indice_fc = np.argmax(P)

# definir listas para graficar el historial de fuerzas y desplazamientos
steps_P = list(range(len(P)))
steps_D = list(range(len(D)))

# crear una lista donde se guardarán los índices seleccionados
selected_points = []

# graficar el historial de fuerzas y desplazamientos
fig, (ax1, ax2) = plt.subplots(nrows = 1, ncols = 2, figsize = (14,8))
line, = ax1.plot(steps_P, P, 'o', color = 'red', ms = 2, picker = True, pickradius = 5)
# line, = ax.plot(steps_D, D, 'o', color='black', ms=2, picker=True, pickradius=5)
line, = ax2.plot(D, P, 'o', color = 'blue', ms = 2, picker = True, pickradius = 5)
crs = mplcursors.cursor([ax1, ax2], hover = True)
crs.connect("add", lambda sel: sel.annotation.set_text('x: {} \ny: {} \nstep: {}'.format(sel.target[0], sel.target[1], sel.target.index)))

fig.tight_layout()


# definir una función que obtenga los puntos al hacer click en la gráfica
def onpick1(event):
    if isinstance(event.artist, Line2D):
        thisline = event.artist
        xdata = thisline.get_xdata()
        ydata = thisline.get_ydata()
        ind = event.ind
        # selected_points.append(int(xdata[ind]))
        selected_points.append(int(ind))
        # print('onpick1 line:', int(xdata[ind]), float(ydata[ind]))
        print('onpick1 line:', int(ind), float(ydata[ind]))

# definir una función que exporte un excel con las curvas separadas
def on_close(event):
    #obtener lista con los índices seleccionados
    lista_sin_repetidos = list(set(selected_points))
    lista_indices = sorted(lista_sin_repetidos)
    
    #exportar a un excel las curvas separadas
    archivo_excel = openpyxl.Workbook()

    for i in range(len(lista_indices)-1):
        k = 0
        a = lista_indices[i]
        b = lista_indices[i+1]

        hoja_trabajo = archivo_excel.create_sheet(f"Hoja{k}")
        
        # Agregar nombres de columna
        hoja_trabajo.cell(row=1, column=1).value = "X [mm]"
        hoja_trabajo.cell(row=1, column=2).value = "Y [tonf]"
        
        # Agregar datos de las dos listas
        hoja_x = D[a:b+1]
        hoja_y = P[a:b+1]
        
        for j in range(len(hoja_x)):
            hoja_trabajo.cell(row=j+2, column=1).value = hoja_x[j]
            hoja_trabajo.cell(row=j+2, column=2).value = hoja_y[j]
            
        k = k + 1
                
    archivo_excel.save('02_Data_excel_01/' + name_data[:-4] +  '.xlsx')
    
    # borrar la primera hoja
    workbook = openpyxl.load_workbook('02_Data_excel_01/' + name_data[:-4] +  '.xlsx')
    hoja_a_borrar = workbook.worksheets[0]
    workbook.remove(hoja_a_borrar)
    workbook.save('02_Data_excel_01/' + name_data[:-4] +  '.xlsx')
    
    print('Índices de los puntos extremos de cada curva:\n', lista_indices)

fig.canvas.mpl_connect('close_event', on_close)

fig.canvas.mpl_connect('pick_event', onpick1)







































